# -*- coding: utf-8 -*-

import os
import yaml
from openpyxl import load_workbook
from osgeo import gdal, ogr, osr
from bs4 import BeautifulSoup
from jinja2 import Environment, FileSystemLoader

JINJA_ENV = Environment(loader=FileSystemLoader('templates'))

TPL_MAP = '''
MAP
NAME "{fc_name}"
SIZE 300 300
UNITS meters

FONTSET "./fonts/fonts.list"

EXTENT {fc_extent}
SHAPEPATH  "{basedir}"

PROJECTION
    "init=epsg:4326"
END

SYMBOL
  TYPE ellipse
  POINTS 1 1 END
  NAME "circle"
END

# Background color for the map canvas -- change as desired
IMAGECOLOR 255 255 255
IMAGEQUALITY 95
IMAGETYPE png

{fc_includes}

END
'''
TPL_LAYER = """
LAYER
    NAME 'land'
    TYPE POLYGON
    DATA 'data'
    STATUS ON
    DUMP true

    # MAXSCALEDENOM 2500000
    # LABELITEM 'RNAME'


    PROJECTION
    "init=epsg:4326"
    END


    METADATA
    'ows_title' ""
    'wms_title' ''
    'ows_srs'  'EPSG:4326 EPSG:3857'
    'wms_srs' 'EPSG:4326 EPSG:3857'
    "gml_include_items"   "all"
    "wms_include_items"   "all"
    "wms_enable_request" "GetMap GetFeatureInfo GetCapabilities"
    END
    CLASS
    end
END
"""
TPL_CLASS = """
CLASS
    STYLE
                
    END
    label
        FONT "simsun"
    end
END
"""

TPL_MAPPROXY = '''
services:
  demo:
  tms:
    use_grid_names: true
    # origin for /tiles service
    origin: 'nw'
  kml:
      use_grid_names: true
  wmts:
  wms:
    md:
      title: MapProxy WMS Proxy
      abstract: This is a minimal MapProxy example.
layers:
  - name: osm
    title: Omniscale OSM WMS - osm.omniscale.net
    sources: [osm_cache]

caches:
  osm_cache:
    grids: [webmercator]
    sources: [osm_wms]

sources:
  osm_wms:
    type: wms
    req:
      # use of this source is only permitted for testing
      url: http://osm.omniscale.net/proxy/service?
      layers: osm

grids:
    webmercator:
        base: GLOBAL_WEBMERCATOR

globals:
'''


def get_mts(afile=None):
    if afile:
        return os.path.getmtime(afile)
    else:
        if os.path.exists('mts.log'):
            return os.path.getatime('mts.log')
        else:
            return 0


def lyr_list(mslug, xls_file):
    wb = load_workbook(filename=xls_file)
    sheet = wb.active

    max_row_num = sheet.max_row

    out_str = []
    for xx in range(1, max_row_num + 1):
        the_cell = 'maplet_' + mslug + '_' + sheet.cell(row=xx, column=1).value
        out_str.append(the_cell)
    return out_str


def get_html_title(html_file):
    uu = BeautifulSoup(open(html_file), 'lxml')
    return uu.title.text


def render_html(tmpl, outfile, **kwargs):
    template = JINJA_ENV.get_template(tmpl)
    output_from_parsed_template = template.render(kwargs)

    # to save the results
    with open(outfile, "w") as fh:
        fh.write(output_from_parsed_template)


def hex2dec(string_num):
    return str(int(string_num.upper(), 16))


def xlsx2dict(xls_file):
    COLOR_INDEX = (
        '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF',
        # 0-4
        '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF',
        # 5-9
        '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',
        # 10-14
        '0000FFFF', '00800000', '00008000', '00000080', '00808000',
        # 15-19
        '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',
        # 20-24
        '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',
        # 25-29
        '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',
        # 30-34
        '0000FFFF', '00800080', '00800000', '00008080', '000000FF',
        # 35-39
        '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',
        # 40-44
        '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',
        # 45-49
        '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',
        # 50-54
        '00969696', '00003366', '00339966', '00003300', '00333300',
        # 55-59
        '00993300', '00993366', '00333399', '00333333',
        'System Foreground', 'System Background'
        # 60-64
    )
    wb = load_workbook(filename=xls_file)
    sheet = wb.active

    max_row_num = sheet.max_row
    max_col_num = sheet.max_column
    out_str = ''
    for xx in range(1, max_row_num + 1):
        # print('x:', xx)
        the_str = ''
        sig = True
        for yy in range(1, max_col_num + 1):
            # print('y:', yy)

            the_cell = sheet.cell(row=xx, column=yy)
            if the_cell and the_cell.value:

                the_cell_value = the_cell.value

                colors = the_cell.fill.fgColor.index
                # print(colors)

                # '00000000' for not filled, 0 for `white`.
                if colors in ['00000000', 0]:
                    pass
                elif len(colors) == 8:
                    a = int(hex2dec(colors[2:4]))
                    b = int(hex2dec(colors[4:6]))
                    c = int(hex2dec(colors[6:8]))
                    the_cell_value = [a, b, c]

                # else:
                #     out_dic[row[0].value] = row[1].value

                if str(the_cell_value).lower() in [
                    'class', 'classitem',
                    'labelitem',
                    'data',
                    'labelminscaledenom',
                    'labelmaxscaledenom',
                    'encoding',
                    'processing',
                ]:
                    the_str = '- ' + the_str

                if sig:
                    the_str = the_str + str(the_cell_value) + ': '
                    sig = False
                else:
                    the_str = the_str + str(the_cell_value)
            else:
                the_str = the_str + '  '
        out_str = out_str + the_str + '\r'

    with open('xx_out.xbj', 'w') as fo:
        fo.write(out_str)
    uu = yaml.load(out_str)
    # from pprint import pprint
    # pprint(uu)

    return uu


def get_epsg_code(img_file, raster=False):
    # print(img_file)
    if os.path.isdir(img_file):
        raster = True
    elif img_file.lower().endswith('.tif'):
        raster = True
    if raster:
        # print(img_file)
        ds = gdal.Open(img_file)
        srs = ds.GetProjection()

        sr2 = osr.SpatialReference()
        sr2.SetFromUserInput(srs)
        # print(sr2.ExportToPrettyWkt())

        return {'epsg_code': '',
                'proj4_code': sr2.ExportToProj4(),
                'geom_type': 'raster'}



    else:
        ds = ogr.Open(img_file)
        lyr = ds.GetLayer(0)
        srs = lyr.GetSpatialRef()

        sr2 = osr.SpatialReference()

        # sr2.SetFromUserInput(srs)
        # epsg_code = srs.GetAttrValue("AUTHORITY", 1)
        # else:
        #     epsg_code = '4326'
        # if epsg_code:
        #     pass
        # else:
        #     epsg_code = '4326'
        geom = None
        idx = 0
        while not geom:
            feat = lyr.GetFeature(idx)
            geom = feat.GetGeometryRef()
            idx = idx + 1
        geom_type = geom.GetGeometryName()

        return {
            'proj4_code': srs.ExportToProj4(),
            # 'epsg_code': '', #epsg_code,
            'geom_type': geom_type}


# ws1=wb.get_sheet_by_name("Sheet1")
if __name__ == '__main__':
    xlsx2dict('meta_poly.xlsx')
